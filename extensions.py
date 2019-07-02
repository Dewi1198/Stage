import openpyxl
import warnings
warnings.filterwarnings("ignore")

extensions = ("default", "pckgdep", "provxml", "resource", "uninstall",
    "manifest", "settingcontent-ms", "automaticdestinations-ms",
    "customdestinations-ms", "log1", "log2", "log3", "log4", "log5", "log6",
    "log7", "log8", "log9", "gm81", "tiger", "catproduct", "kdbx", "note",
    "crdownload", "vdoc", "lock", "game", "vbox", "nitf", "zargo", "safe",
    "vbox-extpack", "html", "ylib", "msproducer", "plsc", "qhcp", "maff",
    "szproj", "blend", "xmpz", "trash", "dired", "123dx", "ipsw", "email",
    "graffle", "evtx", "tocg", "bdmv", "xish", "vbox-prev", "azw2", "icfm",
    "ffivw", "pages", "azw1", "dtsi", "kext", "fsproj", "epub", "aspx", "3dmf",
    "jnlp", "trib", "yookoo", "dbfx", "mdmp", "atml", "flac", "saver",
    "4dindy", "vsdm", "policy", "fasta", "accdt", "mol2", "ingr", "contact",
    "dats", "bufr", "smclvl", "plist", "xisb", "vsdx", "hdmp", "wifi", "vmcz",
    "vala", "wireframe", "face", "rmvb", "vstx", "xmod", "vrml", "dylib",
    "sh3d", "workflow", "xlsb", "plantuml", "spin", "xlsx", "ccitt", "plsk",
    "cpmz", "mpkg", "vlogo", "olsr", "fstick", "jbig", "chml", "tpoly",
    "webloc", "ssif", "ecmt", "docx", "yaml", "mediawiki", "lisp", "escsch",
    "mppz", "scpt", "mpeg", "cats", "3dsx", "scptd", "grdnt", "fb2k",
    "gameproj", "swift", "vbproj", "djvu", "xlsm", "blob", "mobi", "mlraw",
    "adicht", "ampl", "sdts", "psppalette", "lasso", "midi", "aifc",
    "steamstart", "elfo", "ipynb", "iv-vrml", "soar", "spiff", "sldasm",
    "weboogl", "class", "xisf", "sldprt", "numbers", "pipe", "php3", "vsto",
    "rbxl", "tmy2", "flame", "accda", "coffee", "icfe", "odif", "dicom",
    "udiwww", "xlbl", "sha1", "pack", "xspf", "tddd", "ftxt", "unif", "part",
    "vicar", "vmdk", "text", "pncl", "runz", "jpeg", "vstm", "vsqx", "acmb",
    "escpcb", "m2ts", "topc", "fweb", "film", "iptc", "msdl", "iges", "slddrw",
    "dvdproj", "package", "olk14contact", "file", "accdu", "psm1", "schematic",
    "opus", "viff", "jasc", "lang", "eossa", "pblib", "t2flow", "jfif",
    "manager", "wiki", "schdoc", "fsim", "pkpass", "amlx", "uoml", "bmpw",
    "cmrl", "kodu", "m3u8", "hcgs", "dfti", "qrmx", "sndb", "genbank", "grads",
    "naplps", "term", "irtr", "ppsx", "pseg", "dotx", "gmod", "scala",
    "fb2k-component", "vinf", "emaker", "proj", "dbpro", "accft", "sqlite",
    "info", "artx", "torrent", "vssx", "grasp", "rbxm", "miniusf", "aiff",
    "php4", "indd", "kbasic", "rdata", "xosm", "isma", "love", "java", "stuff",
    "onepkg", "uzed", "bpoly", "kfdk", "accdb", "miff", "anim", "wsrc", "objf",
    "hpgl", "qtvr", "oeaccount", "json", "msdvd", "netcdf", "balance",
    "desklink", "xrm-ms", "greenfoot", "hppcl", "adef", "poly", "shar", "font",
    "smpl", "sats", "catdrawing", "sdml", "pict", "ptped", "jsonld", "tiff",
    "rtfd", "themepack", "neis", "grft", "pptx", "cweb", "tria", "attf",
    "sctor", "trif", "enff", "properties", "mime", "rhistory", "msqm", "vssm",
    "catpart", "desktop", "accde", "cals", "docm", "asmx", "proto", "ecms",
    "plugin", "irrmesh", "ply2", "quox", "beam", "par2", "clpi", "vcls",
    "theme", "bdef", "fits", "wmdb", "adson", "clprj", "grib", "ibro", "saif")

newext = []

book = openpyxl.load_workbook("C:\\Users\\Dewiv\\OneDrive\\Documenten\\FileExtensions.xlsx")

s1 = book.get_sheet_by_name('1 - Digital Video')
s2 = book.get_sheet_by_name('2 - Digital Audio')
s3 = book.get_sheet_by_name('3 - Bitmaps')
s4 = book.get_sheet_by_name('4 - Digital Camera (RAW)')
s5 = book.get_sheet_by_name('5 - Document')
s6 = book.get_sheet_by_name('6 - Simple text (ASCII)')
s7 = book.get_sheet_by_name('7 - Spreadsheet')
s8 = book.get_sheet_by_name('8 - MS Office related')
s9 = book.get_sheet_by_name('9 - E-Mail related')
s10 = book.get_sheet_by_name('10 - Internet Related')
s11 = book.get_sheet_by_name('11 - Archive & Compressed')
s12 = book.get_sheet_by_name('12 - Database')
s13 = book.get_sheet_by_name('13 - GIS, GPS, Mapping')
s14 = book.get_sheet_by_name('14 - System files')
s15 = book.get_sheet_by_name('15 - Encoded & Encrypted')


for i in range (2, 655):
   x = s1.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
       newext.append(x)

for i in range (2, 971):
   x = s2.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
    if x not in newext:
       newext.append(x)

for i in range (2, 959):
   x = s3.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
    if x not in newext:
       newext.append(x)

for i in range (2, 74):
   x = s4.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
    if x not in newext:
       newext.append(x)

for i in range (2, 1568):
   x = s5.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
    if x not in newext:
       newext.append(x)

for i in range (2, 513):
   x = s6.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
    if x not in newext:
       newext.append(x)

for i in range (2, 161):
   x = s7.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
    if x not in newext:
       newext.append(x)

for i in range (2, 100):
   x = s8.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
    if x not in newext:
       newext.append(x)

for i in range (2, 74):
   x = s9.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
    if x not in newext:
       newext.append(x)

for i in range (2, 407):
   x = s10.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
    if x not in newext:
       newext.append(x)

for i in range (2, 956):
   x = s11.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
    if x not in newext:
       newext.append(x)

for i in range (2, 468):
   x = s12.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
    if x not in newext:
       newext.append(x)

for i in range (2, 458):
   x = s13.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
    if x not in newext:
       newext.append(x)

for i in range (2, 469):
   x = s14.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
    if x not in newext:
       newext.append(x)

for i in range (2, 465):
   x = s15.cell(row=i, column=4).value
   if (len(str(x)) > 3) and (x not in extensions):
    if x not in newext:
       newext.append(x)


print(*('"{}"'.format(item) for item in newext), sep=", ")
print("Er zijn", len(newext), "nieuwe extensies.")